from __future__ import annotations
import sys, os, json, traceback, time, math, random, collections
from pathlib import Path
from datetime import datetime, timedelta

import cv2
import numpy as np
from PySide6 import QtCore, QtGui, QtWidgets
from ultralytics import YOLO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ================= AJUSTES =================
CONF_THRES       = 0.35
CONF_PERSON      = 0.20
IMGSZ            = 960
FRAME_STRIDE     = 1
PARKED_SECONDS   = 30.0
MOVE_THR_GROUND  = 0.25
IOU_ASSOC_THR    = 0.3

BOX_THICK        = 2
FONT_SCALE       = 0.35
FONT_THICK       = 1
SPLASH_SECS      = 5
ROI_APPLY_SECS   = 3

TRAIL_LEN        = 36          # más puntos para curvas suaves
TRAIL_THICK      = 2
ARROW_THICK      = 2
ARROW_SCALE      = 0.018

MAX_MISSES       = 12

# Tramos persistentes (curvos)
MIN_TRAMO_PIX    = 80          # distancia mínima entre inicio y fin para fijarlo
MIN_PTS_TRAMO    = 10          # puntos mínimos en la trayectoria para que la curva tenga sentido
ANGLE_BIN_DEG    = 25          # agrupación angular para evitar duplicados
MAX_TRAMOS       = 6
SPLINE_RES       = 12          # puntos generados entre cada par (suavidad)
# ===========================================

FINAL_CLASSES = ["vehiculo_menor", "buses_urbanos", "motos", "peatones"]
LABEL_SHORT = {
    "vehiculo_menor": "Veh",
    "buses_urbanos":  "Bus",
    "motos":          "Moto",
    "peatones":       "Peat"
}

CLASS_ROLLUP = {
    "car": "vehiculo_menor", "vehiculo_menor": "vehiculo_menor",
    "bus": "buses_urbanos", "buses_urbanos": "buses_urbanos",
    "motorcycle": "motos", "moto": "motos", "motos": "motos",
    "person": "peatones", "pedestrian": "peatones", "peaton": "peatones", "peatones": "peatones",
    "truck": None, "camion": None, "vehiculos_pesados": None
}

COLOR_MAP = {
    "vehiculo_menor": (0, 255, 255),
    "buses_urbanos":  (255, 255, 0),
    "motos":          (0, 165, 255),
    "peatones":       (255, 0, 255)
}
DEFAULT_COLOR = (60, 170, 255)

# Paleta exclusiva para tramos/direcciones fijas (curvas visibles)
TRAMO_COLORS = [
    ( 50, 220,  50),
    ( 60, 180, 255),
    (255,  80,  80),
    (180,  60, 255),
    (255, 180,  60),
    ( 60, 255, 200)
]

# ---------- UTIL GPU ----------
def device_tuple():
    try:
        import torch
        if torch.cuda.is_available():
            return "cuda:0", 0
    except Exception:
        pass
    return "cpu", "cpu"

def qimage_from_bgr(frame: np.ndarray) -> QtGui.QImage:
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    h, w, ch = rgb.shape
    return QtGui.QImage(rgb.data, w, h, rgb.strides[0], QtGui.QImage.Format.Format_RGB888)

# ============ Loader03_ParticleOrbit ============
from PySide6.QtCore import Qt, QTimer, QRectF, QPointF, QElapsedTimer
from PySide6.QtGui import QColor, QPainter, QBrush

def lerp(a,b,t): return a+(b-a)*t

class Loader03_ParticleOrbit(QtWidgets.QWidget):
    def __init__(self, size=45, color="#00C4FF", speed=1.75, parent=None):
        super().__init__(parent)
        self.size_px=float(size); self.color=QColor(color); self.speed_s=float(speed)
        self.setFixedSize(int(size), int(size))
        self.clock=QElapsedTimer(); self.clock.start()
        self.timer=QTimer(self); self.timer.timeout.connect(self.update); self.timer.start(16)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.angles_deg=[8,36,72,90,144,180,216,252,300,324,335,290,240]
        self.delays=[0,-0.4,-0.9,-0.5,-0.3,-0.2,-0.6,-0.7,-0.1,-0.8,-1.2,-0.5,-0.2]
        self.r_outer=0.5; self.dot_ratio=0.175
        self.center=QPointF(self.size_px/2,self.size_px/2)
        self.dot_r_base=self.size_px*self.dot_ratio/2.0
    def phase01(self, delay=0.0):
        t=self.clock.elapsed()/1000.0; raw=(t+delay)/self.speed_s; return raw-int(raw)
    def _orbit(self, t01):
        tri=1.0-abs(2.0*t01-1.0); r=lerp(0.0,self.r_outer,tri); sc=lerp(0.47,1.0,tri); a=lerp(0.30,1.0,tri)
        return r, sc, a
    def paintEvent(self,_):
        p=QPainter(self); p.setRenderHint(QPainter.Antialiasing, True)
        t=self.clock.elapsed()/1000.0; rot=(t/(self.speed_s*4.0))*2*math.pi
        cosr, sinr=math.cos(rot), math.sin(rot)
        for ang,d in zip(self.angles_deg,self.delays):
            t01=self.phase01(d*self.speed_s); r,sc,a=self._orbit(t01)
            a0=math.radians(ang); vx=math.cos(a0)*self.size_px*r; vy=math.sin(a0)*self.size_px*r
            x=self.size_px/2+(vx*cosr - vy*sinr); y=self.size_px/2+(vx*sinr + vy*cosr)
            R=self.dot_r_base*sc; c=QColor(self.color); c.setAlphaF(a)
            p.setPen(Qt.NoPen); p.setBrush(QBrush(c)); p.drawEllipse(QRectF(x-R,y-R,2*R,2*R))
        p.end()
# =================================================

# ============ Botón Fizzy (con set_label) ==============
from PySide6.QtGui import QPainterPath, QPen, QFont

def hsla(h, s, l, a=1.0):
    c = QtGui.QColor.fromHslF((h % 360) / 360.0, s / 100.0, l / 100.0, a)
    return c

class FizzyParticles:
    def __init__(self, center: QPointF):
        self.center = center
        self.p = []
        self.reset_palette()

    def reset_palette(self):
        self.palette = [hsla(random.randint(180, 340), random.randint(40, 80), random.randint(50, 75)) for _ in range(6)]

    def seed(self, n=64, radius=130, arc_deg=(210, 330)):
        self.p.clear()
        a0, a1 = map(math.radians, arc_deg)
        for i in range(n):
            ang = random.uniform(a0, a1)
            spd = random.uniform(0.15, 0.55)
            r = radius + random.uniform(-8, 8)
            size = random.uniform(6, 11)
            col = self.palette[i % len(self.palette)]
            self.p.append({"ang": ang, "spd": spd, "r": r, "size": size,
                           "alpha": random.uniform(0.55, 0.9), "col": col,
                           "wob": random.uniform(0.6, 1.3), "t": random.uniform(0, 10)})

    def update(self, dt, speed_mult=1.0, compact=1.0):
        for d in self.p:
            d["t"] += dt
            d["ang"] += d["spd"] * dt * speed_mult
            d["r"] += math.sin(d["t"] * d["wob"]) * 0.15 * compact
            if d["ang"] > math.pi * 2: d["ang"] -= math.pi * 2

    def draw(self, qp: QPainter, alpha_scale=1.0, compact=1.0):
        for d in self.p:
            r = d["r"] / compact
            x = self.center.x() + math.cos(d["ang"]) * r
            y = self.center.y() + math.sin(d["ang"]) * r
            a = max(0.0, min(1.0, d["alpha"] * alpha_scale))
            c = QtGui.QColor(d["col"]); c.setAlphaF(a)
            qp.setPen(Qt.NoPen); qp.setBrush(QtGui.QBrush(c))
            s = d["size"] * (0.9 + 0.2 * math.sin(d["t"] * 1.7))
            qp.drawEllipse(QtCore.QRectF(x - s / 2, y - s / 2, s, s))

class FizzyButton(QtWidgets.QWidget):
    clicked = QtCore.Signal()
    def __init__(self, label="Iniciar", parent=None):
        super().__init__(parent)
        self.setMinimumHeight(100)
        self.bg = QtGui.QColor("#24323A")
        self.hover = False
        self.morph = 0.0
        self.timer = QtCore.QTimer(self); self.timer.timeout.connect(self.on_tick); self.timer.start(16)
        self.btn_pos = QtCore.QPointF(0, 0)
        self.w0, self.h0 = 240.0, 72.0
        self.w1, self.h1 = 86.0, 140.0
        self.radius = 22.0
        self._label = label
        self.particles = FizzyParticles(QtCore.QPointF(0, 0))
        self.setup_scene()
        self.setMouseTracking(True)
        self.setMinimumWidth(320)

    def set_label(self, text: str):
        self._label = text
        self.update()

    def setup_scene(self):
        w, h = self.width(), self.height()
        cx, cy = w / 2, h / 2
        self.btn_pos = QtCore.QPointF(cx, cy)
        self.particles.center = QtCore.QPointF(cx, cy + 4)
        self.particles.seed(n=64, radius=145, arc_deg=(210, 330))

    def resizeEvent(self, _): self.setup_scene()
    def mouseMoveEvent(self, e): self.hover = self.point_in_button(e.position()); self.update()
    def mousePressEvent(self, e):
        if e.button() == QtCore.Qt.LeftButton and self.point_in_button(e.position()):
            self.clicked.emit()
    def point_in_button(self, pt: QtCore.QPointF) -> bool:
        w = (1 - self.morph) * self.w0 + self.morph * self.w1
        h = (1 - self.morph) * self.h0 + self.morph * self.h1
        rect = QtCore.QRectF(self.btn_pos.x() - w / 2, self.btn_pos.y() - h / 2, w, h)
        return rect.contains(pt)
    def on_tick(self):
        dt = 0.016
        target = 0.0
        self.morph += (target - self.morph) * min(1.0, 0.12)
        spd = 1.0 + (1.2 if self.hover else 0.0)
        self.particles.update(dt, speed_mult=spd, compact=1.0 + 0.35 * self.morph)
        self.update()

    def path_arrow(self, size=24):
        p = QPainterPath()
        s = size
        p.moveTo(0, s * 0.5)
        p.lineTo(s * 0.72, s * 0.5)
        p.moveTo(s * 0.45, s * 0.28)
        p.lineTo(s * 0.75, s * 0.5)
        p.lineTo(s * 0.45, s * 0.72)
        return p

    def paintEvent(self, _):
        qp = QtGui.QPainter(self); qp.setRenderHint(QtGui.QPainter.Antialiasing, True)
        qp.fillRect(self.rect(), self.bg)
        cx, cy = self.btn_pos.x(), self.btn_pos.y()
        bw = (1 - self.morph) * self.w0 + self.morph * self.w1
        bh = (1 - self.morph) * self.h0 + self.morph * self.h1
        rect = QtCore.QRectF(cx - bw / 2, cy - bh / 2, bw, bh)
        self.particles.center = QtCore.QPointF(cx, cy + 2)
        self.particles.draw(qp, alpha_scale=1.0, compact=1.0)
        qp.setPen(QtCore.Qt.NoPen); qp.setBrush(QtGui.QColor(0,0,0,80))
        qp.drawRoundedRect(rect.adjusted(0,6,0,6), 22*1.2, 22*1.2)
        qp.setBrush(QtGui.QColor("#FFFFFF"))
        qp.setPen(QtGui.QPen(QtGui.QColor(255,255,255), 2))
        qp.drawRoundedRect(rect, 22, 22)
        qp.setPen(QtGui.QPen(QtGui.QColor("#2CC3FF"), 3, QtCore.Qt.SolidLine, QtCore.Qt.RoundCap, QtCore.Qt.RoundJoin))
        qp.save(); qp.translate(rect.left() + 22, rect.center().y() - 12); qp.drawPath(self.path_arrow(26)); qp.restore()
        qp.setPen(QtGui.QPen(QtGui.QColor("#4B5A66")))
        f = QtGui.QFont("Segoe UI", 16, QtGui.QFont.Medium); qp.setFont(f)
        qp.drawText(rect.adjusted(60,0,-16,0), QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft, self._label)
        qp.end()
# ==================================================

# ---------- ROI / Homografía ----------
def load_roi(video_path: Path, W: int, H: int) -> np.ndarray | None:
    f1 = video_path.with_suffix(".roi.json")
    f2 = video_path.parent / "roi.json"
    f = f1 if f1.exists() else (f2 if f2.exists() else None)
    if f is None:
        return None
    try:
        data = json.loads(f.read_text(encoding="utf-8"))
        norm = data.get("normalized", True)
        invert = bool(data.get("invert", False))
        mask = np.zeros((H, W), dtype=np.uint8)
        polys = data.get("polygons") or ([data["points"]] if "points" in data else [])
        if not polys:
            return None
        for pts in polys:
            arr = np.array(pts, dtype=np.float32)
            if norm:
                arr[:, 0] *= W; arr[:, 1] *= H
            cv2.fillPoly(mask, [arr.astype(np.int32)], 255)
        if invert:
            mask = cv2.bitwise_not(mask)
        return mask
    except Exception as e:
        print(f"[ROI] Error al leer {f}: {e}", flush=True)
        return None

def default_quad(W, H):
    return np.array([[0,0],[W-1,0],[W-1,H-1],[0,H-1]], dtype=np.float32)

def four_points_from_roi(mask: np.ndarray) -> np.ndarray | None:
    cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not cnts: return None
    c = max(cnts, key=cv2.contourArea)
    peri = cv2.arcLength(c, True)
    approx = cv2.approxPolyDP(c, 0.02*peri, True)
    if len(approx) != 4: return None
    return approx.reshape(-1,2).astype(np.float32)

def compute_homography(W, H, roi_mask: np.ndarray | None):
    if roi_mask is not None:
        quad = four_points_from_roi(roi_mask)
        if quad is not None:
            def order_pts(pts):
                rect = np.zeros((4,2), dtype="float32")
                s = pts.sum(axis=1); rect[0] = pts[np.argmin(s)]
                rect[2] = pts[np.argmax(s)]
                diff = np.diff(pts, axis=1)
                rect[1] = pts[np.argmin(diff)]
                rect[3] = pts[np.argmax(diff)]
                return rect
            src = order_pts(quad)
            width = max(int(np.linalg.norm(src[1]-src[0])), int(np.linalg.norm(src[2]-src[3])))
            height= max(int(np.linalg.norm(src[3]-src[0])), int(np.linalg.norm(src[2]-src[1])))
            width  = max(100, width); height = max(100, height)
            dst = np.array([[0,0],[width-1,0],[width-1,height-1],[0,height-1]], dtype=np.float32)
            Hmat = cv2.getPerspectiveTransform(src, dst)
            return Hmat, (width, height)
    src = default_quad(W,H)
    dst = np.array([[0,0],[W-1,0],[W-1,H-1],[0,H-1]], dtype=np.float32)
    Hmat = cv2.getPerspectiveTransform(src, dst)
    return Hmat, (W, H)

def warp_point(Hmat: np.ndarray, x: float, y: float) -> tuple[float,float]:
    pt = np.array([[x, y, 1.0]], dtype=np.float32).T
    wp = Hmat @ pt
    w = float(wp[2,0]) if wp[2,0] != 0 else 1e-6
    return float(wp[0,0]/w), float(wp[1,0]/w)

def iou_xyxy(a, b):
    ax1, ay1, ax2, ay2 = a
    bx1, by1, bx2, by2 = b
    ix1, iy1 = max(ax1, bx1), max(ay1, by1)
    ix2, iy2 = min(ax2, bx2), min(ay2, by2)
    iw, ih = max(0.0, ix2 - ix1), max(0.0, iy2 - iy1)
    inter = iw * ih
    if inter <= 0: return 0.0
    area_a = max(0.0,(ax2-ax1)) * max(0.0,(ay2-ay1))
    area_b = max(0.0,(bx2-bx1)) * max(0.0,(by2-by1))
    return inter / max(1e-6, area_a + area_b - inter)

# ---------- Spline / Curvas ----------
def catmull_rom_spline(points: list[tuple[int,int]], res=SPLINE_RES) -> list[tuple[int,int]]:
    """Devuelve puntos suavizados con Catmull-Rom."""
    if len(points) < 4:
        return points[:]
    P = [np.array(p, dtype=np.float32) for p in points]
    out = []
    for i in range(1, len(P)-2):
        p0, p1, p2, p3 = P[i-1], P[i], P[i+1], P[i+2]
        for t in np.linspace(0,1,res,endpoint=False):
            t2, t3 = t*t, t*t*t
            # Catmull-Rom centrípeta
            a = 2*p1
            b = -p0 + p2
            c = 2*p0 - 5*p1 + 4*p2 - p3
            d = -p0 + 3*p1 - 3*p2 + p3
            pt = 0.5*(a + b*t + c*t2 + d*t3)
            out.append((int(pt[0]), int(pt[1])))
    out.append(tuple(P[-2].astype(int)))
    out.append(tuple(P[-1].astype(int)))
    return out

def draw_curve_with_arrow(img, pts, color, thickness, arrow_len_px):
    if len(pts) < 2:
        return
    curve = catmull_rom_spline(pts)
    cv2.polylines(img, [np.array(curve, dtype=np.int32)], False, color, thickness, lineType=cv2.LINE_AA)
    # flecha tangente al final
    p2 = np.array(curve[-1], dtype=np.float32)
    p1 = np.array(curve[-3], dtype=np.float32) if len(curve) >= 3 else np.array(curve[-2], dtype=np.float32)
    v = p2 - p1
    n = np.linalg.norm(v)
    if n < 1e-3: return
    v = (v / n) * max(8.0, arrow_len_px)
    tip = (int(p2[0]), int(p2[1]))
    tail = (int(p2[0]-v[0]), int(p2[1]-v[1]))
    cv2.arrowedLine(img, tail, tip, color, thickness, tipLength=0.35)

def angle_deg_from_vec(vx, vy):
    ang = math.degrees(math.atan2(vy, vx))
    if ang < 0: ang += 360
    return ang

def angle_bin(ang_deg, bin_size=ANGLE_BIN_DEG):
    return int(round(ang_deg / bin_size)) % int(360 / bin_size)

# ---------- Editor de ROI ----------
def edit_roi_opencv(video_path: Path):
    cap = cv2.VideoCapture(str(video_path))
    ok, frame = cap.read(); cap.release()
    if not ok: raise RuntimeError("No pude leer el primer frame para editar ROI.")
    H, W = frame.shape[:2]
    win = "Editar ROI (N:nuevo | U:undo | S:terminar | Q:salir)"
    pts_list: list[list[tuple[int,int]]] = [[]]

    def redraw():
        img = frame.copy()
        for poly in pts_list:
            if len(poly) >= 2:
                cv2.polylines(img, [np.array(poly,np.int32)], False, (0,255,255), 2)
            for p in poly:
                cv2.circle(img, p, 4, (0,200,255), -1)
        cv2.imshow(win, img)

    def on_mouse(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            pts_list[-1].append((x,y))
            redraw()

    cv2.namedWindow(win, cv2.WINDOW_NORMAL)
    cv2.resizeWindow(win, 1200, 700)
    cv2.setMouseCallback(win, on_mouse)
    redraw()

    data = None
    while True:
        key = cv2.waitKey(30) & 0xFF
        if key in (ord('q'), 27):
            break
        elif key in (ord('n'), ord('N')):
            pts_list.append([]); redraw()
        elif key in (ord('u'), ord('U')):
            if pts_list and pts_list[-1]:
                pts_list[-1].pop()
            elif pts_list:
                pts_list.pop()
            if not pts_list: pts_list=[[]]
            redraw()
        elif key in (ord('s'), ord('S')):
            polys = []
            for poly in pts_list:
                if len(poly) >= 3:
                    arr = np.array(poly, dtype=np.float32)
                    arr[:,0] /= W; arr[:,1] /= H
                    polys.append(arr.tolist())
            if not polys:
                cv2.displayOverlay(win, "Dibuja al menos 1 polígono (>=3 puntos).", 2000)
                continue
            data = {"normalized": True, "invert": False, "polygons": polys}
            break

    cv2.destroyWindow(win)
    return data

def save_roi_files(video_path: Path, data: dict):
    out1 = video_path.parent / "roi.json"
    out2 = video_path.with_suffix(".roi.json")
    for out in (out1, out2):
        out.write_text(json.dumps(data, indent=2), encoding="utf-8")
        print(f"[ROI] Guardado {out}", flush=True)

# ---------- Worker (tracks limpios + tramos curvos) ----------
class VideoWorker(QtCore.QObject):
    sig_frame  = QtCore.Signal(QtGui.QImage)
    sig_counts = QtCore.Signal(dict)
    sig_status = QtCore.Signal(str)
    sig_error  = QtCore.Signal(str)
    sig_done   = QtCore.Signal()

    def __init__(self, video_path: str, weights_path: str, use_roi: bool):
        super().__init__()
        self.video_path  = Path(video_path)
        self.weights_path= Path(weights_path)
        self.use_roi     = use_roi
        self._stop       = False

    def stop(self): self._stop = True

    @QtCore.Slot()
    def run(self):
        try:
            dev_str, dev_ultra = device_tuple()
            model = YOLO(str(self.weights_path))
            model.to(dev_str)

            cap = cv2.VideoCapture(str(self.video_path))
            if not cap.isOpened():
                raise RuntimeError(f"No pude abrir el video: {self.video_path}")

            total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
            fps   = float(cap.get(cv2.CAP_PROP_FPS) or 30.0)
            frame_interval = 1.0 / max(1e-6, fps)
            W     = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH) or 0)
            H     = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT) or 0)
            video_dur = timedelta(seconds= total / max(1.0, fps))

            roi_mask = load_roi(self.video_path, W, H) if self.use_roi else None
            Hmat, _ = compute_homography(W, H, roi_mask)

            tracks: dict[int, dict] = {}
            next_id = 1
            counts = {k: 0 for k in FINAL_CLASSES}

            # Tramos persistentes curvos
            tramos = []  # cada tramo: dict(curve_pts=[(x,y)...], color=(b,g,r), bin=int)
            used_bins = set()

            wall_start = time.time()

            while True:
                if self._stop:
                    self.sig_status.emit("Detenido por el usuario.")
                    break

                elapsed = time.time() - wall_start
                desired_idx = int(elapsed * fps)
                current_pos = int(cap.get(cv2.CAP_PROP_POS_FRAMES))
                if desired_idx - current_pos > 1:
                    cap.set(cv2.CAP_PROP_POS_FRAMES, max(0, desired_idx - 1))

                ok, frame = cap.read()
                if not ok: break

                res = model.predict(source=[frame], imgsz=IMGSZ,
                                    conf=min(CONF_THRES, CONF_PERSON),
                                    device=dev_ultra, verbose=False)[0]

                dets = []
                if res.boxes is not None and len(res.boxes) > 0:
                    xyxy = res.boxes.xyxy.cpu().numpy()
                    clss = res.boxes.cls.cpu().numpy().astype(int)
                    conf = res.boxes.conf.cpu().numpy()
                    for bb, c, cf in zip(xyxy, clss, conf):
                        cname = str(res.names.get(int(c), int(c))).lower()
                        need_conf = CONF_PERSON if ("person" in cname or "peat" in cname) else CONF_THRES
                        if cf < need_conf: continue
                        roll = CLASS_ROLLUP.get(cname, cname)
                        if roll is None or roll not in FINAL_CLASSES: continue
                        x1, y1, x2, y2 = [int(v) for v in bb.tolist()]
                        cx, cy = int((x1+x2)/2), int((y1+y2)/2)
                        if self.use_roi and roi_mask is not None:
                            if cx < 0 or cy < 0 or cx >= W or cy >= H or roi_mask[cy, cx] == 0:
                                continue
                        dets.append(((x1,y1,x2,y2), roll, float(cf), (cx,cy)))

                # Asociación IoU
                pairs=[]
                for tid,t in tracks.items():
                    for j,(bb,cn,cf,(cx,cy)) in enumerate(dets):
                        pairs.append((iou_xyxy(t["xyxy"], bb), tid, j))
                pairs=[p for p in pairs if p[0] >= IOU_ASSOC_THR]
                pairs.sort(reverse=True)

                used_tids=set(); used_js=set()
                for _, tid, j in pairs:
                    if tid in used_tids or j in used_js: continue
                    bb, roll, cf, (cx,cy) = dets[j]
                    gx, gy = warp_point(Hmat, cx, cy)
                    gprev = tracks[tid]["center_ground"]
                    gdist = float(np.hypot(gx - gprev[0], gy - gprev[1]))
                    if gdist <= MOVE_THR_GROUND:
                        tracks[tid]["still_s"] += frame_interval
                    else:
                        tracks[tid]["still_s"] = 0.0
                    tracks[tid].update(dict(
                        xyxy=bb, center_img=(cx,cy), center_ground=(gx,gy),
                        cls=roll, misses=0
                    ))
                    tracks[tid]["trail"].append((cx, cy))
                    used_tids.add(tid); used_js.add(j)

                # nuevos tracks
                for j,(bb,roll,cf,(cx,cy)) in enumerate(dets):
                    if j in used_js: continue
                    gx, gy = warp_point(Hmat, cx, cy)
                    tracks[next_id] = dict(
                        xyxy=bb, center_img=(cx,cy), center_ground=(gx,gy),
                        still_s=0.0, cls=roll, counted=False,
                        trail=collections.deque(maxlen=TRAIL_LEN),
                        misses=0
                    )
                    tracks[next_id]["trail"].append((cx, cy))
                    next_id += 1

                # incrementar misses y limpiar
                to_del=[]
                for tid,t in tracks.items():
                    if tid not in used_tids:
                        t["misses"] = t.get("misses",0) + 1
                    if t.get("misses",0) > MAX_MISSES:
                        to_del.append(tid)
                for tid in to_del:
                    tracks.pop(tid, None)

                # Conteo 1xID no estacionado
                for tid,t in list(tracks.items()):
                    roll = t["cls"]
                    if roll not in FINAL_CLASSES: 
                        continue
                    is_vehicle = roll in ("vehiculo_menor","buses_urbanos","motos")
                    parked = is_vehicle and (t["still_s"] >= PARKED_SECONDS)
                    if not parked and not t["counted"]:
                        counts[roll] += 1
                        t["counted"] = True

                # Fijar tramos CURVOS persistentes (una vez por dirección)
                if len(tramos) < MAX_TRAMOS:
                    for t in tracks.values():
                        trail = list(t.get("trail", []))
                        if len(trail) >= MIN_PTS_TRAMO:
                            p0 = np.array(trail[0], dtype=np.float32)
                            pN = np.array(trail[-1], dtype=np.float32)
                            v  = pN - p0
                            if np.linalg.norm(v) >= MIN_TRAMO_PIX:
                                ang = angle_deg_from_vec(v[0], v[1])
                                bin_id = angle_bin(ang, ANGLE_BIN_DEG)
                                if bin_id not in used_bins:
                                    color = TRAMO_COLORS[len(tramos) % len(TRAMO_COLORS)]
                                    # guardamos la curva suavizada completa
                                    curve_pts = catmull_rom_spline(trail, res=SPLINE_RES)
                                    tramos.append(dict(curve=curve_pts, color=color, bin=bin_id))
                                    used_bins.add(bin_id)
                                    if len(tramos) >= MAX_TRAMOS:
                                        break

                # Dibujo
                vis = frame.copy()

                # 1) Curvas persistentes (línea + flecha)
                arrow_len_px = int(min(W, H) * ARROW_SCALE)
                for tr in tramos:
                    draw_curve_with_arrow(vis, tr["curve"], tr["color"], ARROW_THICK, arrow_len_px)

                # 2) Trail temporal detrás de cajas
                for t in tracks.values():
                    roll=t["cls"]
                    if roll not in FINAL_CLASSES:
                        continue
                    color = COLOR_MAP.get(roll, DEFAULT_COLOR)
                    trail_pts = list(t.get("trail", []))
                    if len(trail_pts) >= 3:
                        curve_tmp = catmull_rom_spline(trail_pts, res=max(6, SPLINE_RES//2))
                        cv2.polylines(vis, [np.array(curve_tmp, dtype=np.int32)], False, color, 1, lineType=cv2.LINE_AA)

                # 3) Cajas + etiquetas pequeñas
                for t in tracks.values():
                    roll=t["cls"]; 
                    if roll not in FINAL_CLASSES: 
                        continue
                    color = COLOR_MAP.get(roll, DEFAULT_COLOR)
                    x1,y1,x2,y2 = [int(v) for v in t["xyxy"]]
                    tag = LABEL_SHORT.get(roll, roll)
                    cv2.rectangle(vis,(x1,y1),(x2,y2),color,BOX_THICK)
                    (tw,th),_ = cv2.getTextSize(tag, cv2.FONT_HERSHEY_SIMPLEX,FONT_SCALE,FONT_THICK)
                    pad = 4
                    cv2.rectangle(vis,(x1,max(0,y1-th-2*pad)),(x1+tw+2*pad,y1),color,-1)
                    cv2.putText(vis,tag,(x1+pad,y1-4),cv2.FONT_HERSHEY_SIMPLEX,FONT_SCALE,(0,0,0),FONT_THICK,cv2.LINE_AA)

                if self.use_roi and roi_mask is not None:
                    overlay = vis.copy()
                    overlay[roi_mask==0] = (overlay[roi_mask==0]*0.4).astype(np.uint8)
                    vis = overlay

                self.sig_frame.emit(qimage_from_bgr(vis))
                self.sig_counts.emit(counts.copy())

                current_frame_num = int(cap.get(cv2.CAP_PROP_POS_FRAMES))
                elapsed_td = timedelta(seconds=max(0.0, time.time() - wall_start))
                remaining_frames = max(0, total - current_frame_num)
                eta = timedelta(seconds=remaining_frames * frame_interval)
                self.sig_status.emit(f"Procesando frame {current_frame_num}/{total} | dur={video_dur} | t={elapsed_td} | ETA~{eta}")

                target_time = current_frame_num * frame_interval
                ahead = target_time - (time.time() - wall_start)
                if ahead > 0:
                    time.sleep(min(ahead, 0.010))

            cap.release()
            self.sig_status.emit("Finalizado.")
            self.sig_done.emit()

        except Exception as e:
            tb = traceback.format_exc()
            self.sig_error.emit(f"{e}\n{tb}")
            self.sig_done.emit()

# ---------- Helpers de pesos ----------
def auto_find_best(start_dir: Path) -> Path | None:
    candidates = []
    for base in [start_dir, start_dir.parent, Path.cwd()]:
        runs = base / "runs" / "detect"
        if runs.exists():
            candidates += list(runs.rglob("weights/best.pt"))
    if (start_dir/"best.pt").exists():
        candidates.append(start_dir/"best.pt")
    if (Path.cwd()/"best.pt").exists():
        candidates.append(Path.cwd()/"best.pt")
    if not candidates: return None
    candidates = sorted(set(candidates), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]

# ---------- Overlay spinner centrado ----------
class CenterSpinner(QtWidgets.QWidget):
    def __init__(self, parent: QtWidgets.QWidget):
        super().__init__(parent)
        self.setAttribute(QtCore.Qt.WA_TransparentForMouseEvents, True)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint | QtCore.Qt.Tool)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)
        self.loader = Loader03_ParticleOrbit(size=90, color="#00C4FF", speed=1.75, parent=self)
        self.label = QtWidgets.QLabel("Procesando…", self)
        self.label.setStyleSheet("color:#EAF6FF; font: 16px 'Segoe UI';")
        lay = QtWidgets.QVBoxLayout(self); lay.setContentsMargins(0,0,0,0)
        lay.addStretch(1)
        inner = QtWidgets.QVBoxLayout(); inner.addWidget(self.loader, 0, QtCore.Qt.AlignCenter)
        inner.addWidget(self.label, 0, QtCore.Qt.AlignCenter)
        lay.addLayout(inner); lay.addStretch(1)

    def resizeEvent(self, _):
        if self.parent():
            self.setGeometry(self.parent().rect())

    def show_for_seconds(self, secs=3):
        self.resizeEvent(None)
        self.show()
        QtCore.QTimer.singleShot(int(secs*1000), self.hide)

# ---------- Ventana principal ----------
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Conteo Vehicular — YOLO (GPU)")
        self._thread: QtCore.QThread | None = None
        self._worker: VideoWorker | None = None

        self.video_path: Path | None = None
        self.dataset_dir: Path | None = None
        self.weights_path: Path | None = auto_find_best(Path.cwd())
        self.counts = {k: 0 for k in FINAL_CLASSES}
        self.roi_mask_preview: np.ndarray | None = None

        QtWidgets.QApplication.setStyle("Fusion")
        dark = QtGui.QPalette()
        dark.setColor(QtGui.QPalette.Window, QtGui.QColor(37,37,38))
        dark.setColor(QtGui.QPalette.WindowText, QtCore.Qt.white)
        dark.setColor(QtGui.QPalette.Base, QtGui.QColor(30,30,30))
        dark.setColor(QtGui.QPalette.Text, QtCore.Qt.white)
        dark.setColor(QtGui.QPalette.Button, QtGui.QColor(45,45,48))
        dark.setColor(QtGui.QPalette.ButtonText, QtCore.Qt.white)
        self.setPalette(dark)

        central = QtWidgets.QWidget(self); self.setCentralWidget(central)
        root = QtWidgets.QVBoxLayout(central)

        # Top controls
        top = QtWidgets.QHBoxLayout()
        self.btn_video   = QtWidgets.QPushButton("Cargar Video")
        self.btn_dataset = QtWidgets.QPushButton("Cargar Dataset")
        self.btn_weights = QtWidgets.QPushButton("Elegir best.pt")
        self.cb_roi      = QtWidgets.QCheckBox("Filtrar por ROI"); self.cb_roi.setChecked(True); self.cb_roi.setStyleSheet("color:#DDD;")
        self.btn_roi     = QtWidgets.QPushButton("Editar ROI")
        for b, col in [(self.btn_video, "#007ACC"), (self.btn_dataset, "#007ACC"),
                       (self.btn_weights, "#6C757D"), (self.btn_roi, "#FF9800")]:
            b.setStyleSheet(f"background:{col};color:white;")
            top.addWidget(b)
        top.addWidget(self.cb_roi)
        root.addLayout(top)

        # Viewer
        self.lbl_view = QtWidgets.QLabel("Sin video")
        self.lbl_view.setAlignment(QtCore.Qt.AlignCenter)
        self.lbl_view.setMinimumHeight(460)
        self.lbl_view.setStyleSheet("background:#1E1E1E;border:1px solid #333;")
        root.addWidget(self.lbl_view, 3)

        # Tabla
        self.table = QtWidgets.QTableWidget(0,2)
        self.table.setHorizontalHeaderLabels(["Categoría","Conteo único"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        root.addWidget(self.table, 1)

        # Botones inferiores
        bottom = QtWidgets.QHBoxLayout()
        self.fizzy_start = FizzyButton("Iniciar")
        self.btn_stop    = QtWidgets.QPushButton("Detener")
        self.btn_export  = QtWidgets.QPushButton("Exportar Excel")
        for b, col in [(self.btn_stop, "#DC3545"), (self.btn_export, "#17A2B8")]:
            b.setStyleSheet(f"background:{col};color:white;")
        bottom.addWidget(self.fizzy_start, 1)
        bottom.addWidget(self.btn_stop)
        bottom.addWidget(self.btn_export)
        root.addLayout(bottom)

        self.lbl_status = QtWidgets.QLabel("Listo.")
        root.addWidget(self.lbl_status)

        # Señales
        self.btn_video.clicked.connect(self.pick_video)
        self.btn_dataset.clicked.connect(self.pick_dataset)
        self.btn_weights.clicked.connect(self.pick_weights)
        self.fizzy_start.clicked.connect(self.start_detection)
        self.btn_stop.clicked.connect(self.stop_detection)
        self.btn_export.clicked.connect(self.export_excel)
        self.btn_roi.clicked.connect(self.edit_roi)

        self._refresh_counts()
        if self.weights_path:
            self.lbl_status.setText(f"Pesos por defecto: {self.weights_path}")
        else:
            self.lbl_status.setText("Sin pesos. Carga dataset o best.pt.")

    def show_frame_preview_with_roi(self):
        if not self.video_path: return
        cap = cv2.VideoCapture(str(self.video_path))
        ok, frame = cap.read(); cap.release()
        if not ok: return
        H, W = frame.shape[:2]
        mask = load_roi(self.video_path, W, H)
        if mask is None:
            self.roi_mask_preview = None
            img = frame
        else:
            self.roi_mask_preview = mask
            overlay = frame.copy()
            overlay[mask==0] = (overlay[mask==0]*0.4).astype(np.uint8)
            img = overlay
        self.lbl_view.setPixmap(QtGui.QPixmap.fromImage(qimage_from_bgr(img)).scaled(
            self.lbl_view.size(), QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation))

    def pick_video(self):
        p, _ = QtWidgets.QFileDialog.getOpenFileName(self,"Selecciona video","", "Video (*.mp4 *.mov *.avi)")
        if not p: return
        self.video_path = Path(p)
        self.lbl_status.setText(f"Video: {p}")
        try:
            data = edit_roi_opencv(self.video_path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error al editar ROI", str(e)); data=None

        if data is not None:
            res = QtWidgets.QMessageBox.question(
                self, "Guardar ROI", "¿Deseas guardar este ROI?",
                QtWidgets.QMessageBox.StandardButton.Ok | QtWidgets.QMessageBox.StandardButton.Cancel,
                QtWidgets.QMessageBox.StandardButton.Ok
            )
            if res == QtWidgets.QMessageBox.StandardButton.Ok:
                blocker = CenterSpinner(self); blocker.show_for_seconds(ROI_APPLY_SECS)
                save_roi_files(self.video_path, data)
                self.lbl_status.setText("ROI guardado. Aplicado al video.")
            else:
                self.lbl_status.setText("ROI cancelado por el usuario.")
        self.show_frame_preview_with_roi()

    def pick_dataset(self):
        d = QtWidgets.QFileDialog.getExistingDirectory(self,"Selecciona carpeta del dataset")
        if not d: return
        self.dataset_dir = Path(d)
        if not (self.dataset_dir/"dataset.yaml").exists():
            QtWidgets.QMessageBox.warning(self,"Dataset inválido","No se encontró dataset.yaml.")
            return
        candidate = auto_find_best(self.dataset_dir)
        if candidate:
            self.weights_path=candidate; self.lbl_status.setText(f"Dataset OK. Pesos: {candidate}")
        else:
            self.lbl_status.setText("Dataset OK. Usa 'Elegir best.pt'.")

    def pick_weights(self):
        p,_=QtWidgets.QFileDialog.getOpenFileName(self,"Selecciona best.pt","", "Pesos (*.pt)")
        if p: self.weights_path=Path(p); self.lbl_status.setText(f"Pesos: {p}")

    def edit_roi(self):
        if not self.video_path:
            QtWidgets.QMessageBox.information(self, "Falta video", "Carga primero un video para editar el ROI.")
            return
        try:
            data = edit_roi_opencv(self.video_path)
            if data is None:
                self.lbl_status.setText("ROI no guardado."); return
            res = QtWidgets.QMessageBox.question(
                self, "Guardar ROI", "¿Deseas guardar este ROI?",
                QtWidgets.QMessageBox.StandardButton.Ok | QtWidgets.QMessageBox.StandardButton.Cancel,
                QtWidgets.QMessageBox.StandardButton.Ok
            )
            if res == QtWidgets.QMessageBox.StandardButton.Ok:
                blocker = CenterSpinner(self); blocker.show_for_seconds(ROI_APPLY_SECS)
                save_roi_files(self.video_path, data)
                self.lbl_status.setText("ROI guardado. Aplicado al video.")
            else:
                self.lbl_status.setText("ROI cancelado por el usuario.")
            self.show_frame_preview_with_roi()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error al editar ROI", str(e))

    def _refresh_counts(self):
        self.table.setRowCount(len(self.counts))
        for i,(k,v) in enumerate(sorted(self.counts.items())):
            self.table.setItem(i,0,QtWidgets.QTableWidgetItem(k))
            self.table.setItem(i,1,QtWidgets.QTableWidgetItem(str(v)))

    def _enable_controls(self, enabled: bool):
        for w in [self.btn_video,self.btn_dataset,self.btn_weights,self.btn_roi,self.cb_roi,self.fizzy_start,self.btn_export]:
            w.setEnabled(enabled)
        self.btn_stop.setEnabled(not enabled)

    def start_detection(self):
        if not self.video_path:
            QtWidgets.QMessageBox.warning(self,"Falta video","Selecciona un video."); return
        if not self.weights_path or not self.weights_path.exists():
            QtWidgets.QMessageBox.warning(self,"Faltan pesos","Selecciona best.pt del entrenamiento."); return
        self.counts = {k: 0 for k in FINAL_CLASSES}
        self._refresh_counts()

        self._thread = QtCore.QThread(self)
        self._worker = VideoWorker(str(self.video_path), str(self.weights_path), self.cb_roi.isChecked())
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.sig_frame.connect(self.on_frame)
        self._worker.sig_counts.connect(self.on_counts)
        self._worker.sig_status.connect(self.on_status)
        self._worker.sig_error.connect(self.on_error)
        self._worker.sig_done.connect(self.on_done)
        self._thread.finished.connect(self._cleanup_thread)

        self._thread.start()
        self._enable_controls(False)
        self.fizzy_start.set_label("Analizando datos…")
        self.lbl_status.setText("Iniciando…")

    def stop_detection(self):
        if self._worker:
            self._worker.stop()
            self.fizzy_start.set_label("Detenido")
            self.lbl_status.setText("Deteniendo…")

    @QtCore.Slot()
    def _cleanup_thread(self):
        self._thread=None; self._worker=None
        self._enable_controls(True)

    @QtCore.Slot(QtGui.QImage)
    def on_frame(self, img: QtGui.QImage):
        pix = QtGui.QPixmap.fromImage(img)
        self.lbl_view.setPixmap(pix.scaled(self.lbl_view.size(),
            QtCore.Qt.AspectRatioMode.KeepAspectRatio,
            QtCore.Qt.TransformationMode.SmoothTransformation))

    @QtCore.Slot(dict)
    def on_counts(self, counts: dict):
        self.counts = counts; self._refresh_counts()

    @QtCore.Slot(str)
    def on_status(self, text: str):
        self.lbl_status.setText(text); print(text, flush=True)

    @QtCore.Slot(str)
    def on_error(self, msg: str):
        print(msg, flush=True); QtWidgets.QMessageBox.critical(self,"Error",msg)

    @QtCore.Slot()
    def on_done(self):
        self.lbl_status.setText("Listo.")
        self.fizzy_start.set_label("Terminado")
        self._enable_controls(True)

    def export_excel(self):
        if not self.counts:
            QtWidgets.QMessageBox.information(self,"Sin datos","No hay conteos para exportar."); return
        save_path,_=QtWidgets.QFileDialog.getSaveFileName(self,"Guardar Excel","conteo.xlsx","Excel (*.xlsx)")
        if not save_path: return

        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Conteo"
        bold=Font(bold=True,color="FFFFFF")
        fill=PatternFill("solid",fgColor="4F81BD")
        border=Border(left=Side(style="thin"),right=Side(style="thin"),
                      top=Side(style="thin"),bottom=Side(style="thin"))
        now=datetime.now().strftime("%Y-%m-%d %H:%M")

        ws.merge_cells("A1:B1"); ws["A1"]=f"Conteo vehicular — {now}"
        ws["A1"].font=Font(bold=True,size=14); ws["A1"].alignment=Alignment(horizontal="center")

        ws["A2"],ws["B2"]="Categoría","Conteo"
        for c in ["A2","B2"]:
            ws[c].font=bold; ws[c].fill=fill; ws[c].alignment=Alignment(horizontal="center")

        r=3; maxlenA=len("Categoría"); maxlenB=len("Conteo")
        for k,v in sorted(self.counts.items()):
            ws[f"A{r}"]=k; ws[f"B{r}"]=v
            ws[f"A{r}"].border=border; ws[f"B{r}"].border=border
            ws[f"A{r}"].alignment=Alignment(vertical="center")
            ws[f"B{r}"].alignment=Alignment(horizontal="center",vertical="center")
            maxlenA=max(maxlenA,len(str(k))); maxlenB=max(maxlenB,len(str(v))); r+=1

        ws.column_dimensions["A"].width=max(12, min(40, int(maxlenA*1.2)))
        ws.column_dimensions["B"].width=max(10, min(20, int(maxlenB*1.2)+2))

        wb.save(save_path)
        self.lbl_status.setText(f"Excel exportado: {save_path}")

# ---------- Splash ----------
class AppWithSplash(QtWidgets.QApplication):
    def __init__(self, argv):
        super().__init__(argv)
        self.setStyle("Fusion")
        self._splash = QtWidgets.QWidget()
        self._splash.setWindowFlags(QtCore.Qt.FramelessWindowHint | QtCore.Qt.SplashScreen)
        self._splash.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)
        lay = QtWidgets.QVBoxLayout(self._splash); lay.setContentsMargins(0, 0, 0, 0)
        panel = QtWidgets.QWidget(); panel.setStyleSheet("background:rgba(22,28,34,230); border-radius:16px;")
        panel_lay = QtWidgets.QVBoxLayout(panel); panel_lay.setContentsMargins(30,30,30,30)
        loader = Loader03_ParticleOrbit(size=110, color="#00C4FF", speed=1.6)
        title = QtWidgets.QLabel("ULTRAANALYTIC — YOLO"); title.setStyleSheet("color:#EAF6FF; font: 20px 'Segoe UI Semibold';")
        sub   = QtWidgets.QLabel("Inicializando…");         sub.setStyleSheet("color:#8ACBFF; font: 14px 'Segoe UI';")
        panel_lay.addWidget(loader, 0, QtCore.Qt.AlignCenter)
        panel_lay.addSpacing(10)
        panel_lay.addWidget(title, 0, QtCore.Qt.AlignCenter)
        panel_lay.addWidget(sub,   0, QtCore.Qt.AlignCenter)
        lay.addStretch(1); lay.addWidget(panel, 0, QtCore.Qt.AlignCenter); lay.addStretch(1)
        self._splash.resize(520, 320); self._splash.show()
        QtCore.QTimer.singleShot(SPLASH_SECS*1000, self._splash.close)

def main():
    def excepthook(exc_type, exc_value, exc_tb):
        tb="".join(traceback.format_exception(exc_type,exc_value,exc_tb))
        print(tb, flush=True)
        QtWidgets.QMessageBox.critical(None,"Error no manejado",tb)
    sys.excepthook=excepthook

    app = AppWithSplash(sys.argv)
    win=MainWindow()
    win.showMaximized()
    sys.exit(app.exec())

if __name__=="__main__":
    main()
